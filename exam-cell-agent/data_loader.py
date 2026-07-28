"""
data_loader.py — Loads student/course enrolment data.
Derives: branch, is_arrear, year, credits, roll_range per dept.
"""
import csv
import json
import re
from config import sem_to_year

try:
    import pandas as pd
except Exception:
    pd = None


def _parse_branch(reg_no: str) -> str:
    """Extract branch code from reg_no, e.g. '722824104001' -> '104' or 'CS'."""
    match = re.search(r"[A-Z]{2,6}", reg_no)
    if match:
        return match.group(0)
    # Check 12-digit format e.g., 722823104001 (branch code digits 6..9)
    clean = str(reg_no).strip()
    if clean.startswith("7228") and len(clean) >= 9:
        return clean[6:9]
    return "UNKNOWN"


def _parse_reg_no_info(reg_no: str) -> dict:
    """
    Extract college code and batch year from reg_no.
    Standard reg_no format: '7228YY...' (e.g. 722823104001)
      - College code: '7228'
      - Batch code: '23', '24', '25', '26'
    Maps batch to regular semester:
      '26' -> Sem 1 (1st Year)
      '25' -> Sem 3 (2nd Year)
      '24' -> Sem 5 (3rd Year)
      '23' -> Sem 7 (4th Year)
    """
    clean = str(reg_no).strip()
    batch = "25"
    if clean.startswith("7228") and len(clean) >= 6:
        batch = clean[4:6]
    elif clean.startswith("202") and len(clean) >= 6:
        batch = clean[2:4]
    elif len(clean) >= 2 and clean[:2].isdigit():
        batch = clean[:2]

    batch_map = {
        "26": 1,
        "25": 3,
        "24": 5,
        "23": 7,
    }
    regular_sem = batch_map.get(batch, None)
    return {
        "batch": batch,
        "regular_sem": regular_sem,
    }


def _roll_range(reg_nos: list[str]) -> str:
    """Return 'MIN–MAX' roll range string from a list of reg_nos."""
    if not reg_nos:
        return ""
    sorted_rolls = sorted(reg_nos)
    return f"{sorted_rolls[0]}–{sorted_rolls[-1]}"


def load_students(source) -> list[dict]:
    """
    Load student enrolment data from CSV, JSON, or Excel (.xlsx/.xls).

    Columns expected for CSV/JSON: name, reg_no, course_code, course_name, semester
    Excel workbook expected sheets:
      1. 'Arrear Details' (Columns: Sl. No., Branch, Sem, Code, Register Number)
      2. 'Regular Courses' (Columns: Branch, Sem, Code)

    Derives:
      branch       — from reg_no or sheet
      year         — from semester (sem 1-2 → yr 1, 3-4 → yr 2, etc.)
      is_arrear    — enforced via 7228YY batch mapping (sem != regular_sem)
      credits      — from column or default 3
    """
    if isinstance(source, str) and (source.endswith(".xlsx") or source.endswith(".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(source, data_only=True)

        enrolments = []
        students_map = {}

        # 1. Load Arrear Details
        ws_arr = wb["Arrear Details"] if "Arrear Details" in wb.sheetnames else wb.worksheets[0]
        arr_rows = list(ws_arr.iter_rows(values_only=True))

        header_idx = 0
        for i, r in enumerate(arr_rows[:5]):
            row_strs = [str(x).strip().lower() for x in r if x is not None]
            if "code" in row_strs or "register number" in row_strs or "sem" in row_strs:
                header_idx = i
                break

        h_arr = [str(c).strip().lower() for c in arr_rows[header_idx]]
        b_col = h_arr.index("branch") if "branch" in h_arr else 1
        s_col = h_arr.index("sem") if "sem" in h_arr else 2
        c_col = h_arr.index("code") if "code" in h_arr else 3
        r_col = h_arr.index("register number") if "register number" in h_arr else 4

        for row in arr_rows[header_idx + 1:]:
            if not row or all(v is None for v in row):
                continue
            branch = str(row[b_col]).strip() if len(row) > b_col and row[b_col] else "UNKNOWN"
            try:
                sem = int(row[s_col]) if len(row) > s_col and row[s_col] else 1
            except (ValueError, TypeError):
                sem = 1
            code = str(row[c_col]).strip() if len(row) > c_col and row[c_col] else ""
            raw_reg = str(row[r_col]).split(".")[0].strip() if len(row) > r_col and row[r_col] else ""
            if not raw_reg or not code:
                continue

            batch = raw_reg[4:6] if len(raw_reg) >= 6 else "25"
            if raw_reg not in students_map:
                students_map[raw_reg] = {"branch": branch, "batch": batch, "max_sem": sem}
            else:
                students_map[raw_reg]["max_sem"] = max(students_map[raw_reg]["max_sem"], sem)

            enrolments.append({
                "name": f"Student {raw_reg}",
                "reg_no": raw_reg,
                "branch": branch,
                "semester": sem,
                "year": sem_to_year(sem),
                "course_code": code,
                "course_name": code,
                "credits": 3,
                "is_arrear": True,
            })

        # 2. Load Regular Courses (if available)
        if "Regular Courses" in wb.sheetnames:
            ws_reg = wb["Regular Courses"]
            reg_rows = list(ws_reg.iter_rows(values_only=True))
            reg_h_idx = None
            for i, r in enumerate(reg_rows[:10]):
                if not r:
                    continue
                r_strs = [str(x).strip().lower() for x in r if x is not None]
                if "branch" in r_strs and "code" in r_strs:
                    reg_h_idx = i
                    break

            reg_courses = {}
            if reg_h_idx is not None:
                h_reg = [str(c).strip().lower() if c else "" for c in reg_rows[reg_h_idx]]
                rb_col = h_reg.index("branch") if "branch" in h_reg else 2
                rs_col = h_reg.index("sem") if "sem" in h_reg else 3
                rc_col = h_reg.index("code") if "code" in h_reg else 4

                for row in reg_rows[reg_h_idx + 1:]:
                    if not row or len(row) <= max(rb_col, rs_col, rc_col):
                        continue
                    b = str(row[rb_col]).strip() if row[rb_col] else ""
                    try:
                        s = int(row[rs_col]) if row[rs_col] else None
                    except (ValueError, TypeError):
                        s = None
                    c = str(row[rc_col]).strip() if row[rc_col] else ""
                    if b and s and c:
                        reg_courses.setdefault((b, s), []).append(c)

            # Expand regular enrolments for each student
            for raw_reg, s_info in students_map.items():
                branch = s_info["branch"]
                batch = s_info["batch"]
                if batch == "25":
                    reg_sem = 3
                elif batch == "24":
                    reg_sem = 5
                elif batch == "23":
                    reg_sem = 7
                else:
                    reg_sem = s_info["max_sem"] + 1 if (s_info["max_sem"] % 2 == 0) else s_info["max_sem"] + 2

                c_list = reg_courses.get((branch, reg_sem), [])
                for c_code in c_list:
                    enrolments.append({
                        "name": f"Student {raw_reg}",
                        "reg_no": raw_reg,
                        "branch": branch,
                        "semester": reg_sem,
                        "year": sem_to_year(reg_sem),
                        "course_code": c_code,
                        "course_name": c_code,
                        "credits": 3,
                        "is_arrear": False,
                    })

        return enrolments

    if isinstance(source, list):
        raw_records = source
    elif isinstance(source, str) and source.endswith(".json"):
        with open(source, encoding="utf-8") as f:
            raw_records = json.load(f)
    elif isinstance(source, str) and (source.endswith(".csv") or not source.endswith(".xlsx")):
        with open(source, mode="r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            raw_records = list(reader)
    elif pd is not None:
        df = pd.read_csv(source)
        raw_records = df.to_dict(orient="records")
    else:
        raw_records = []

    column_alias = {
        "register_number": "reg_no",
        "register_no": "reg_no",
        "reg_number": "reg_no",
        "roll_no": "reg_no",
        "code": "course_code",
        "subject_code": "course_code",
        "sem": "semester",
        "dept": "branch",
        "department": "branch",
    }

    student_max_sem = {}
    for r in raw_records:
        norm_r = {}
        for k, v in r.items():
            norm_k = str(k).strip().lower().replace(" ", "_")
            norm_r[column_alias.get(norm_k, norm_k)] = v
        reg = str(norm_r.get("reg_no", "")).strip()
        try:
            sem = int(norm_r.get("semester", 1))
        except (ValueError, TypeError):
            sem = 1
        if reg:
            student_max_sem[reg] = max(student_max_sem.get(reg, 1), sem)

    enrolments = []
    for r in raw_records:
        norm_r = {}
        for k, v in r.items():
            norm_k = str(k).strip().lower().replace(" ", "_")
            norm_r[column_alias.get(norm_k, norm_k)] = v

        reg_no = str(norm_r.get("reg_no", "")).strip()
        course_code = str(norm_r.get("course_code", "")).strip()
        if not reg_no or not course_code:
            continue

        try:
            sem = int(norm_r.get("semester", 1))
        except (ValueError, TypeError):
            sem = 1

        name = str(norm_r.get("name", f"Student {reg_no}"))
        course_name = str(norm_r.get("course_name", course_code))
        branch = str(norm_r.get("branch", "")).strip() or _parse_branch(reg_no)

        try:
            credits_val = int(norm_r.get("credits", 3))
        except (ValueError, TypeError):
            credits_val = 3

        if "is_arrear" in norm_r and norm_r["is_arrear"] is not None and str(norm_r["is_arrear"]).strip() != "":
            val = str(norm_r["is_arrear"]).strip().lower()
            is_arr = val in ("true", "1", "yes", "y", "t")
        else:
            info = _parse_reg_no_info(reg_no)
            if info["regular_sem"] is not None:
                is_arr = (sem != info["regular_sem"])
            else:
                is_arr = (sem < student_max_sem.get(reg_no, sem))

        enrolments.append({
            "name": name,
            "reg_no": reg_no,
            "branch": branch,
            "semester": sem,
            "year": sem_to_year(sem),
            "course_code": course_code,
            "course_name": course_name,
            "credits": credits_val,
            "is_arrear": is_arr,
        })

    return enrolments


def build_dept_roll_ranges(enrolments: list[dict]) -> dict[str, dict]:
    """
    Build dept-wise roll ranges per semester.

    Returns:
        {branch: {semester: "24CS001–24CS320"}}
    """
    from collections import defaultdict
    dept_sem_rolls: dict[tuple, list] = defaultdict(list)
    for row in enrolments:
        dept_sem_rolls[(row["branch"], row["semester"])].append(row["reg_no"])

    result: dict[str, dict] = defaultdict(dict)
    for (branch, sem), rolls in dept_sem_rolls.items():
        result[branch][sem] = _roll_range(rolls)
    return dict(result)

